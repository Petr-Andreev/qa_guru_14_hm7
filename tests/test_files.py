import csv
import zipfile
from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader


def test_read_xlsx_file():
    with ZipFile('resource/my_archive.zip', 'r') as zip_file:
        with zip_file.open("file_example_XLSX_50.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            cell_value = sheet.cell(row=2, column=5).value
            print(f"Содержимое ячейки в строке {2} и столбце {5}: {cell_value}")
            name = 'United States'
            assert name in cell_value, f"Cтрана: {name} отсутствует в списке"


def test_csv():
    with zipfile.ZipFile('resource/my_archive.zip', 'r') as zip_file:
        with zip_file.open("import_empl_csv.csv") as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')  # читаем содержимое файла и декодируем его если в файле есть символы не из английского алфавита
            csvreader = list(csv.reader(content.splitlines()))  # читаем содержимое файла и преобразуем его в список
            second_row = csvreader[1]  # получаем вторую строку
            result_list = second_row[0].split(';')
            model = 'CS4164U'
            manufacturer = "B4COM"
            assert result_list[2] == manufacturer, f"Производитель: {manufacturer
            } отсутствует в списке {result_list}"  # проверка значения элемента в первом столбце второй строки
            assert result_list[3] == model, f"Модель: {model
            } отсутствует в списке {model}"  # проверка значения элемента во втором столбце второй строки


def test_read_pdf_file():
    with ZipFile('resource/my_archive.zip', 'r') as zip_file:
        with zip_file.open("python_testing_with_pytest.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[1]  # получаем первую страницу
            text = page.extract_text()  # извлекаем текст из первой страницы
            assert 'Stable Diffusion' in text